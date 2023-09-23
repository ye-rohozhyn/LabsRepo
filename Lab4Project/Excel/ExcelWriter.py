import openpyxl


def write_excel(filename, data, header=None, sheet_name="Sheet"):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        if header is not None:
            sheet.append(header)

        for row in data:
            sheet.append(row)

        workbook.save(filename)
        print(f"Таблиця у форматі XLSX створена та збережена у файлі: {filename}")
    except Exception as e:
        print(f"Виникла помилка під час створення файлу Excel: {str(e)}")


def add_sheet(filename, data, header=None, sheet_name: str = "Sheet"):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.create_sheet(title=sheet_name)

        if header is not None:
            sheet.append(header)

        for row in data:
            sheet.append(row)

        workbook.save(filename)
        print(f"Сторінка у форматі XLSX створена та збережена у файлі: {filename} за назвою: {sheet_name}")
    except Exception as e:
        print(f"Виникла помилка під час створення сторінки для файлу Excel: {str(e)}")
